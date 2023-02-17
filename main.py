import datetime
import tkinter as tk
import tkinter.ttk as ttk 
import math

import requests
import openpyxl
import pandas as pd

from bs4 import BeautifulSoup
from threading import Thread
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment


class App():
	def __init__(self, size_win, title):
		self.size_win = size_win
		self.window = tk.Tk()
		self.title = title
		self.start = True

		self.element = {
			'label': {},
			'radio_button': {},
			'button': {},
			'input': {},
			'scale': {},
			'progressbar': {},
			'check_button': {}
		}

		self.window_mainloop()

	# Запуск потока
	def start_script(self):
		if self.start:
			self.start = False
			thread = Thread(target=self.main_parsing)
			thread.start()
		else:
			if 'error' in self.element['label'].keys():
				self.element_delete('label', 'error')
			self.element_text('error', 'Скрипт запущен, ожидайте результата!')
	
	# Запуск парсинга
	def main_parsing(self):
		self.club_parsing_list = []
		self.club_list = {}
		self.table_club = []

		if 'progressbar' in self.element['progressbar'].keys():
			self.element_delete('progressbar', 'progressbar')
		

		self.element_progressbar('progressbar', 100)
		if self.element['check_button']['all_club'].get() == 1:
			self.get_all_table()
		else:
			if 'error' in self.element['label'].keys():
				self.element_delete('label', 'error') 
			if 'https://24score.pro' not in self.element['input']['url_input'].get():
				self.element_text('error', 'Введите ссылку!')
				return 

			self.get_url_club(self.element['input']['url_input'].get())

		self.element['progressbar']['progressbar']['value'] = self.element['progressbar']['progressbar']['value'] + 20
		self.parsing_club()
		self.check_match()
		self.save_file()

		self.start = True
	
	# Получения всех ссылок на все таблицы
	def get_all_table(self):
		r = requests.get('https://24score.pro/football')
		url_table_list = BeautifulSoup(r.text, 'html.parser').find('ul', class_='champlist').find_all('li')

		for url_table in url_table_list:
			url = f"https://24score.pro{url_table.find('a').get('href')}"
			self.get_url_club(url)

		if self.element['radio_button']['select_type'].get() == 0:
			self.file_name = f'ТекущаяДата({datetime.datetime.now().strftime("%d_%m_%Y")})_ПоВсем_Чет'
		else:
			self.file_name = f'ТекущаяДата({datetime.datetime.now().strftime("%d_%m_%Y")})_ПоВсем_Нечет'

	# Получения ссылок на клубы
	def get_url_club(self, url):
		with requests.Session() as sess:
			r = sess.get(url)
			for str_ in r.text.split('\n'):
				if 'data: {"data_key" : "' in str_:
					id_ = str_.replace('data: {"data_key" : "', '').replace('"},', '').strip()
			resp = sess.get(f'https://24score.pro/backend/load_page_data.php?data_key={id_}')


		title_url = BeautifulSoup(r.text, 'html.parser').find('ul', class_='champlist').find('li', class_='selected').find('a').get('title').title().replace('-', '').replace(' ', '')
		if self.element['radio_button']['select_type'].get() == 0:
			self.file_name = f'ТекущаяДата({datetime.datetime.now().strftime("%d_%m_%Y")})_{title_url}_Чет'
		else:
			self.file_name = f'ТекущаяДата({datetime.datetime.now().strftime("%d_%m_%Y")})_{title_url}_Нечет'



		search_bs = BeautifulSoup(resp.text, 'html.parser')

		table = search_bs.find('table', class_='standings report evenodd').find_all('td')

		for column in table:
			if column.find('a') != None:
				url = f"https://24score.pro{column.find('a').get('href')}"
				self.club_parsing_list.append([url, column.find('a').get_text(strip=True)])
	
	# Парсинг клуба
	def parsing_club(self):
		brogres_count = 60 / len(self.club_parsing_list)

		for club_url in self.club_parsing_list:
			r = requests.get(club_url[0])
			search_bs = BeautifulSoup(r.text, 'html.parser')

			match = {
				'url': club_url[0],
				'matche_played': [],
				'matche_future': []
			}

			table = search_bs.find('table', class_='datatable oddeven').find_all('tr')

			for column in table:
				if column.find('td', class_='date') == None or column.find('td', class_='date').text.strip() == '':
					continue

				data_ = column.find('td', class_='date').text

				team_first = column.find_all('td', class_='team')[0].get_text(strip=True)
				team_second = column.find_all('td', class_='team')[1].get_text(strip=True)

				if column.find('a', {'target': '_blank'}).text.strip() != '':
					match['matche_played'].append({
						'data': data_,
						'team_first': team_first,
						'team_second': team_second,
						'score': column.find('a', {'target': '_blank'}).text
					})
				else:
					match['matche_future'].append({
						'data': data_,
						'team_first': team_first,
						'team_second': team_second,
					})

			self.club_list[club_url[1]] = match
			self.element['progressbar']['progressbar']['value'] = self.element['progressbar']['progressbar']['value'] + brogres_count
	
	# Сортировка клубов
	def check_match(self):
		value_played = int(self.element['scale']['matche_played'].get())
		value_future = int(self.element['scale']['matche_future'].get())
		value_type = self.element['radio_button']['select_type'].get()

		for club in self.club_list:
			name_ = club
			match_played = self.club_list[club]['matche_played'][:value_played]
			matche_future = list(reversed(self.club_list[club]['matche_future']))[:value_future]
			matche_list = []
			matche_count = 0


			for matche_f in matche_future:
				data_ = [int(datetime.datetime.now().year), int(datetime.datetime.now().month),  int(datetime.datetime.now().day)]
				if datetime.date(int(matche_f['data'].split('.')[2]), int(matche_f['data'].split('.')[1]), int(matche_f['data'].split('.')[0])) >= datetime.date(data_[0], data_[1], data_[2]):
					matche_list.append(matche_f)

			
			for match_ in match_played:
				score = match_['score']

				first_goal = int(score.split(':')[0])
				second_goal = int(score.split(':')[1])

				if (first_goal % 2 == 0 and club == match_['team_first']) or (second_goal % 2 == 0 and club == match_['team_second']):
					if value_type == 0:
						matche_count += 1
				else:
					if value_type == 1:
						matche_count += 1

			if matche_count == value_played:
				self.table_club.append({
						'club': name_,
						'matche': matche_list,
						'url': self.club_list[club]['url']
					})
	
	# Сохраняем результат
	def save_file(self):
		max_data = datetime.date(int(datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=3))).year), int(datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=3))).month), int(datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=3))).day))
		club_list = []

		for club in self.table_club:
			club_name = club['club']
			club_matche = club['matche']
			matche_data = []

			for matche in club_matche:
				data = matche['data'].split('.')
				matche_data.append(matche['data'])
				if datetime.date(int(data[2]), int(data[1]), int(data[0])) > max_data:
					max_data = datetime.date(int(data[2]), int(data[1]), int(data[0]))

			club_list.append({
					'name': club_name,
					'data': matche_data,
					'url': club['url']
				})

		data_list = pd.date_range(
			min(datetime.date(int(datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=3))).year), int(datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=3))).month), int(datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=3))).day)), max_data),
			max(datetime.date(int(datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=3))).year), int(datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=3))).month), int(datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=3))).day)), max_data)
		).strftime('%d.%m.%Y').tolist()

		wb = openpyxl.Workbook()
		list_ = wb.active

		wb = openpyxl.Workbook()
		list_ = wb.active
		list_["C2"] = 'Возврат'
		list_["C3"] = 'Проигрыш'
		list_["C4"] = 'Выигрыш'

		list_['B2'].fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type = "solid")
		list_['B3'].fill = PatternFill(start_color="00FF0000", end_color="00FF0000", fill_type = "solid")
		list_['B4'].fill = PatternFill(start_color="00008000", end_color="00008000", fill_type = "solid")
		list_['B9'] = 'Команда'

		border_style = Side(border_style="thin", color="000000")
		color_club_name = ''
		club_file_list = {}

		if self.element['radio_button']['select_type'].get() == 0:
			list_['B6'] = 'Сыграно четных игр'
			list_['B6'].font = Font(size=12)
			list_['B7'] = int(self.element['scale']['matche_played'].get())
			list_['B7'].font = Font(size=12)

			list_['B8'].fill = PatternFill(start_color="00FF0000", end_color="00FF0000", fill_type = "solid")
			list_['C8'] = 'Будущие четные игры'
			list_['C8'].font = Font(size=12)

			list_['F8'] = 'Ставь на Индивидуальный тотал Чет/нет'
			list_['F8'].font = Font(size=12, bold=True)
			color_club_name = PatternFill(start_color="00FFCC99", end_color="00FFCC99", fill_type = "solid")
		else:
			list_['B6'] = 'Сыграно нечетных игр'
			list_['B6'].font = Font(size=12)
			list_['B7'] = int(self.element['scale']['matche_played'].get())

			list_['B8'].fill = PatternFill(start_color="0000CCFF", end_color="0000CCFF", fill_type = "solid")
			list_['C8'] = 'Будущие нечетные игры'
			list_['B8'].font = Font(size=12)

			list_['F8'] = 'Ставь на Индивидуальный тотал Чет/да'
			list_['F8'].font = Font(size=12, bold=True)
			color_club_name = PatternFill(start_color="0099CCFF", end_color="0099CCFF", fill_type = "solid")

		club_index = 10
		for club in club_list:
			list_[f'B{club_index}'] = '=HYPERLINK("{}", "{}")'.format(club["url"], club["name"])
			list_[f'B{club_index}'].font = Font(size=12)
			list_[f'B{club_index}'].fill = color_club_name

			if club_list.index(club) != len(club_list) - 1:
				list_[f'B{club_index + 1}'].fill = color_club_name

			club_file_list[club['name']] = {
				'index': club_index,
				'data': club['data']
			}

			club_index += 2

		data_cell_list = {}
		data_index = 3
		for data in data_list:
			list_.cell(9, data_index).value = data
			list_.cell(9, data_index).border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)
			list_.cell(9, data_index).font = Font(size=12)
			data_cell_list[data] = data_index
			data_index += 1

		for club in club_file_list:
			for data in club_file_list[club]['data']:
				list_.cell(club_file_list[club]['index'], data_cell_list[data]).value = 'Да'
				list_.cell(club_file_list[club]['index'], data_cell_list[data]).font = Font(size=12)
				list_.cell(club_file_list[club]['index'], data_cell_list[data]).alignment = Alignment(horizontal='center')


		wb.save(f'{self.file_name}.xlsx')
		self.element['progressbar']['progressbar']['value'] = 100
		self.element_text('error', 'Готово!')


	# Создать текст
	def element_text(self, teg, text):
		self.element['label'][teg] = tk.Label(self.window, text=text)
		self.element['label'][teg].pack()
	
	# Создать радио-кнопку
	def element_radio_button(self, teg, type_list):
		self.element['radio_button'][teg] = tk.IntVar()
		self.element['radio_button'][teg].set(0)

		for type_ in type_list:
			type_button = tk.Radiobutton(
				self.window,
				text=type_,
				variable=self.element['radio_button'][teg], 
				value=type_list.index(type_)
			)
			type_button.pack()
	
	# Создать чек-кнопку
	def element_check_button(self, teg, name, function):
		self.element['check_button'][teg] = tk.IntVar()
		self.element['check_button'][teg].set(0)
		check_button = tk.Checkbutton(self.window, text=name, variable=self.element['check_button'][teg], command=function)
		check_button.pack()

	# Создать кнопку		
	def element_button(self, teg, text, function):
		self.element['button'][teg] = tk.Button(self.window, text=text, command=function)
		self.element['button'][teg].pack()
	
	# Создать элемент ввода
	def element_input(self, teg):
		message = tk.StringVar()
		self.element['input'][teg] = tk.Entry(textvariable=message)
		self.element['input'][teg].pack()
	
	# Создать ползунок 
	def element_scale(self, teg, min_value, max_value, value):
		self.element['scale'][teg] = tk.Scale(self.window, showvalue=4, from_=min_value, to=max_value, orient=tk.HORIZONTAL)
		self.element['scale'][teg].set(value)
		self.element['scale'][teg].pack()
	
	# Создать прогрессбар
	def element_progressbar(self, teg, length):
		self.element['progressbar'][teg] = ttk.Progressbar(self.window, orient=tk.HORIZONTAL, length=length, mode='determinate')
		self.element['progressbar'][teg].pack()
	
	# Удалить элемент
	def element_delete(self, element, teg):
		self.element[element][teg].pack_forget()

	# setting
	def element_setting(self):
		self.element_text('title', 'Введите URL таблицы')
		self.element_input('url_input') 
		self.element['input']['url_input'].bind('<Key>', lambda event: self.element['check_button']['all_club'].set(0))
		

		self.element_check_button('all_club', 'Общая таблица', lambda: self.element['input']['url_input'].delete(0, tk.END))

		self.element_text('matche_future', 'Число будущих матчей')
		self.element_scale('matche_future', 1, 7, 4)

		self.element_text('matche_played', 'Число сыгранных матчей')
		self.element_scale('matche_played', 1, 15, 4)

		self.element_radio_button('select_type', ['ЧЁТ', 'НЕЧЁТ'])

		self.element_button('start_script', 'Получить', lambda: self.start_script())

	# mainloop
	def window_mainloop(self):
		self.window.title(self.title)
		self.window.geometry(f'{self.size_win[0]}x{self.size_win[1]}')

		self.element_setting()

		self.window.mainloop()
		


if __name__ == '__main__':
	App([400, 500], 'Парсинг')
